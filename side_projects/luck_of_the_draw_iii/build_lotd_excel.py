"""Download Luck of the Draw III, import it into the isolated lotd schema, and export Excel."""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from tqdm import tqdm

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
	sys.path.insert(0, str(PROJECT_ROOT))

from backend.database import engine


PARQUET_URLS = [
	"https://huggingface.co/datasets/refugee-law-lab/luck-of-the-draw-iii"
	"/resolve/refs%2Fconvert%2Fparquet/default/train/0000.parquet",
	"https://huggingface.co/datasets/refugee-law-lab/luck-of-the-draw-iii"
	"/resolve/refs%2Fconvert%2Fparquet/default/train/0001.parquet",
]
PROJECT_DIR = Path(__file__).resolve().parent
CACHE_DIR = PROJECT_DIR / "cache"
OUTPUT_DIR = PROJECT_DIR / "output"
OUTPUT_PATH = OUTPUT_DIR / "luck_of_the_draw_iii.xlsx"
HF_HEADERS = {"User-Agent": "AI-CaseLibrary/LotD-importer"}

CASE_COLUMNS = [
	"IMM_NUMBER", "YEAR", "NAME", "DATE_FILED", "CITY_FILED", "NATURE", "CLASS", "TRACK",
	"DOC_COUNT", "SOURCE_URL", "SCRAPED_TIMESTAMP",
]
DOCKET_COLUMNS = ["IMM_NUMBER", "RE_NO", "DOCNO", "DOC_DT", "RECORDED_ENTRY", "SOURCE_URL"]
COL_WIDTHS = {
	"IMM_NUMBER": 18, "YEAR": 8, "NAME": 45, "DATE_FILED": 14, "CITY_FILED": 16,
	"NATURE": 42, "CLASS": 20, "TRACK": 25, "DOC_COUNT": 12, "SCRAPED_TIMESTAMP": 22,
	"RE_NO": 9, "DOCNO": 9, "DOC_DT": 14, "RECORDED_ENTRY": 80, "SOURCE_URL": 55,
}
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description=__doc__)
	parser.add_argument("--skip-download", action="store_true")
	parser.add_argument("--skip-excel", action="store_true")
	parser.add_argument("--batch-size", type=int, default=5_000)
	return parser.parse_args()


def download(url: str, destination: Path) -> None:
	if destination.exists() and destination.stat().st_size:
		print(f"  [cached] {destination.name}")
		return
	print(f"  Downloading {destination.name}...")
	with requests.get(url, headers=HF_HEADERS, stream=True, timeout=180) as response:
		response.raise_for_status()
		total = int(response.headers.get("Content-Length", 0))
		with destination.open("wb") as file_obj, tqdm(
			total=total, unit="B", unit_scale=True, unit_divisor=1024, leave=False
		) as progress:
			for chunk in response.iter_content(1 << 20):
				if chunk:
					file_obj.write(chunk)
					progress.update(len(chunk))


def parse_documents(value: Any) -> list[dict[str, Any]]:
	if isinstance(value, list):
		return [item for item in value if isinstance(item, dict)]
	if hasattr(value, "tolist"):
		converted = value.tolist()
		if isinstance(converted, list):
			return [item for item in converted if isinstance(item, dict)]
	if not isinstance(value, str) or not value.strip():
		return []
	try:
		parsed = json.loads(value)
	except json.JSONDecodeError:
		return []
	return [item for item in parsed if isinstance(item, dict)] if isinstance(parsed, list) else []


def normalized_cases(raw: pd.DataFrame) -> pd.DataFrame:
	cases = raw[
		["citation", "year", "name", "date_filed", "city_filed", "nature", "class", "track", "documents", "source_url", "scraped_timestamp"]
	].copy()
	cases["DOC_COUNT"] = cases["documents"].map(lambda value: len(parse_documents(value)))
	cases = cases.drop(columns=["documents"]).rename(columns={
		"citation": "IMM_NUMBER", "year": "YEAR", "name": "NAME", "date_filed": "DATE_FILED",
		"city_filed": "CITY_FILED", "nature": "NATURE", "class": "CLASS", "track": "TRACK",
		"source_url": "SOURCE_URL", "scraped_timestamp": "SCRAPED_TIMESTAMP",
	})
	cases = cases.dropna(subset=["IMM_NUMBER"]).drop_duplicates("IMM_NUMBER", keep="last")
	return cases[CASE_COLUMNS]


def iter_docket_records(raw: pd.DataFrame) -> Iterable[dict[str, Any]]:
	for row in raw[["citation", "documents", "source_url"]].itertuples(index=False):
		imm_number, documents, source_url = row
		if pd.isna(imm_number):
			continue
		for document in parse_documents(documents):
			yield {
				"IMM_NUMBER": str(imm_number), "RE_NO": document.get("RE_NO"), "DOCNO": document.get("DOCNO"),
				"DOC_DT": document.get("DOC_DT"), "RECORDED_ENTRY": document.get("RECORDED_ENTRY"),
				"SOURCE_URL": source_url,
			}


def chunks(rows: Iterable[dict[str, Any]], size: int) -> Iterable[list[dict[str, Any]]]:
	batch: list[dict[str, Any]] = []
	for row in rows:
		batch.append(row)
		if len(batch) == size:
			yield batch
			batch = []
	if batch:
		yield batch


def create_isolated_schema(connection) -> None:
	connection.execute(text("CREATE SCHEMA IF NOT EXISTS lotd"))
	connection.execute(text("""
		CREATE TABLE IF NOT EXISTS lotd.cases (
			imm_number TEXT PRIMARY KEY, year TEXT, name TEXT, date_filed TEXT, city_filed TEXT,
			nature TEXT, class TEXT, track TEXT, doc_count INTEGER NOT NULL, source_url TEXT,
			scraped_timestamp TEXT
		)
	"""))
	connection.execute(text("""
		CREATE TABLE IF NOT EXISTS lotd.dockets (
			id BIGSERIAL PRIMARY KEY, imm_number TEXT NOT NULL REFERENCES lotd.cases(imm_number) ON DELETE CASCADE,
			re_no TEXT, docno TEXT, doc_dt TEXT, recorded_entry TEXT, source_url TEXT
		)
	"""))
	connection.execute(text("CREATE INDEX IF NOT EXISTS ix_lotd_dockets_imm_number ON lotd.dockets (imm_number)"))
	connection.execute(text("CREATE INDEX IF NOT EXISTS ix_lotd_dockets_doc_dt ON lotd.dockets (doc_dt)"))
	connection.execute(text("CREATE INDEX IF NOT EXISTS ix_lotd_dockets_re_no ON lotd.dockets (re_no)"))
	connection.execute(text("""
		CREATE OR REPLACE VIEW lotd.access_cases AS
		SELECT
			imm_number,
			year,
			name,
			date_filed,
			city_filed,
			nature,
			class,
			track,
			doc_count,
			source_url,
			scraped_timestamp
		FROM lotd.cases
		ORDER BY imm_number
	"""))
	connection.execute(text("""
		CREATE OR REPLACE VIEW lotd.access_docket_summary AS
		SELECT
			c.imm_number,
			c.name,
			c.year,
			c.date_filed,
			c.city_filed,
			c.nature,
			c.class,
			c.track,
			c.doc_count,
			MIN(NULLIF(d.doc_dt, '')) AS first_doc_dt,
			MAX(NULLIF(d.doc_dt, '')) AS last_doc_dt,
			MAX(COALESCE(NULLIF(d.re_no, ''), '0')::numeric) AS max_re_no,
			COUNT(*) FILTER (WHERE COALESCE(NULLIF(d.docno, ''), '') <> '') AS docket_rows_with_docno
		FROM lotd.cases c
		LEFT JOIN lotd.dockets d ON d.imm_number = c.imm_number
		GROUP BY
			c.imm_number, c.name, c.year, c.date_filed, c.city_filed,
			c.nature, c.class, c.track, c.doc_count
		ORDER BY c.imm_number
	"""))
	connection.execute(text("""
		CREATE OR REPLACE VIEW lotd.access_dockets AS
		SELECT
			id,
			imm_number,
			re_no,
			docno,
			doc_dt,
			recorded_entry,
			source_url
		FROM lotd.dockets
		ORDER BY imm_number, re_no NULLS LAST, id
	"""))


def import_database(cases: pd.DataFrame, dockets: Iterable[dict[str, Any]], batch_size: int) -> int:
	case_rows = cases.where(pd.notna(cases), None).to_dict("records")
	with engine.begin() as connection:
		create_isolated_schema(connection)
		connection.execute(text("TRUNCATE lotd.dockets, lotd.cases"))
		connection.execute(text("""
			INSERT INTO lotd.cases (imm_number, year, name, date_filed, city_filed, nature, class, track, doc_count, source_url, scraped_timestamp)
			VALUES (:IMM_NUMBER, :YEAR, :NAME, :DATE_FILED, :CITY_FILED, :NATURE, :CLASS, :TRACK, :DOC_COUNT, :SOURCE_URL, :SCRAPED_TIMESTAMP)
		"""), case_rows)
		insert_dockets = text("""
			INSERT INTO lotd.dockets (imm_number, re_no, docno, doc_dt, recorded_entry, source_url)
			VALUES (:IMM_NUMBER, :RE_NO, :DOCNO, :DOC_DT, :RECORDED_ENTRY, :SOURCE_URL)
		""")
		docket_count = 0
		for batch in tqdm(chunks(dockets, batch_size), desc="  Import dockets", unit="batch"):
			connection.execute(insert_dockets, batch)
			docket_count += len(batch)
	return docket_count


def configure_sheet(sheet, headers: list[str], row_count: int) -> None:
	for column_index, header in enumerate(headers, 1):
		sheet.column_dimensions[get_column_letter(column_index)].width = COL_WIDTHS.get(header, 18)
	sheet.freeze_panes = "A2"
	sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count + 1}"


def append_row(sheet, values: Iterable[Any], headers: list[str], row_number: int) -> None:
	row = []
	for column_index, value in enumerate(values, 1):
		cell = WriteOnlyCell(sheet, value=value)
		cell.font = BODY_FONT
		cell.alignment = WRAP_ALIGNMENT if headers[column_index - 1] == "RECORDED_ENTRY" else TOP_ALIGNMENT
		if row_number % 2 == 0:
			cell.fill = PatternFill("solid", fgColor="DCE6F1")
		row.append(cell)
	sheet.append(row)


def append_header(sheet, headers: list[str]) -> None:
	row = []
	for header in headers:
		cell = WriteOnlyCell(sheet, value=header)
		cell.font = HDR_FONT
		cell.fill = HDR_FILL
		cell.alignment = Alignment(horizontal="center", vertical="center")
		row.append(cell)
	sheet.append(row)


def export_excel() -> tuple[int, int]:
	with engine.connect() as connection:
		case_count = connection.execute(text("SELECT COUNT(*) FROM lotd.cases")).scalar_one()
		docket_count = connection.execute(text("SELECT COUNT(*) FROM lotd.dockets")).scalar_one()
		workbook = Workbook(write_only=True)
		case_sheet = workbook.create_sheet("Cases")
		append_header(case_sheet, CASE_COLUMNS)
		for row_number, row in enumerate(connection.execute(text("""
			SELECT imm_number, year, name, date_filed, city_filed, nature, class, track, doc_count, source_url, scraped_timestamp
			FROM lotd.cases ORDER BY imm_number
		""")), 2):
			append_row(case_sheet, row, CASE_COLUMNS, row_number)
		configure_sheet(case_sheet, CASE_COLUMNS, case_count)

		docket_sheet = workbook.create_sheet("Dockets")
		append_header(docket_sheet, DOCKET_COLUMNS)
		for row_number, row in enumerate(connection.execution_options(stream_results=True).execute(text("""
			SELECT imm_number, re_no, docno, doc_dt, recorded_entry, source_url
			FROM lotd.dockets ORDER BY imm_number, re_no NULLS LAST, id
		""")), 2):
			append_row(docket_sheet, row, DOCKET_COLUMNS, row_number)
		configure_sheet(docket_sheet, DOCKET_COLUMNS, docket_count)

		OUTPUT_DIR.mkdir(exist_ok=True)
		workbook.save(OUTPUT_PATH)
	return case_count, docket_count


def main() -> None:
	args = parse_args()
	if args.batch_size < 1:
		raise SystemExit("--batch-size must be at least 1")
	CACHE_DIR.mkdir(exist_ok=True)
	if not args.skip_download:
		print("=== Step 1: Download parquet files ===")
		for url in PARQUET_URLS:
			download(url, CACHE_DIR / url.rsplit("/", 1)[-1])

	parquet_files = sorted(CACHE_DIR.glob("*.parquet"))
	if len(parquet_files) != len(PARQUET_URLS):
		raise SystemExit("Both parquet files must be present in the cache.")
	print("=== Step 2: Load and normalize parquet ===")
	raw = pd.concat([pd.read_parquet(path) for path in parquet_files], ignore_index=True)
	cases = normalized_cases(raw)
	print(f"  Raw records: {len(raw):,}; cases: {len(cases):,}")

	print("=== Step 3: Import isolated SQL tables ===")
	docket_count = import_database(cases, iter_docket_records(raw), args.batch_size)
	print(f"  Dockets imported: {docket_count:,}")
	if args.skip_excel:
		return
	print("=== Step 4: Write Excel workbook ===")
	started = time.perf_counter()
	case_count, docket_count = export_excel()
	print(f"  Workbook: {OUTPUT_PATH.resolve()}")
	print(f"  Cases: {case_count:,}; Dockets: {docket_count:,}; elapsed: {time.perf_counter() - started:.1f}s")


if __name__ == "__main__":
	main()